import pandas as pd
from openpyxl import load_workbook
from json import dumps

datapath = r"C:\Users\Nicole\Desktop\python_projects\DE Guide 2024\CDP_2018.xlsx"
# Load Excel workbook

wb = load_workbook(datapath)

# Loop through all sheets
for sheet in wb.worksheets:

    # Find the number of rows and columns in the sheet
    rows = sheet.max_row
    columns = sheet.max_column

    # List to store all rows as dictionaries
    lst = []

    # Iterate over rows and columns to extract data
    for i in range(1, rows):
        row = {}
        for j in range(1, columns):
            column_name = sheet.cell(row=1, column=j)
            row_data = sheet.cell(row=i+1, column=j)

            row.update(
                {
                    column_name.value: row_data.value
                }
            )
        lst.append(row)

    # Convert extracted data into JSON format
    json_data = dumps(lst, indent=4)

    # Print the JSON data
    print(json_data)
